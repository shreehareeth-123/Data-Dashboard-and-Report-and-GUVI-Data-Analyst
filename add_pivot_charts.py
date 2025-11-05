# add_pivot_charts.py
# Loads Sales_Dashboard.xlsx, creates summary tables and inserts charts into a new sheet.

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, Reference, BarChart, LineChart
from openpyxl.chart.label import DataLabelList

# Read data
df = pd.read_excel('Sales_Dashboard.xlsx')

# Prepare summaries
sales_by_product = df.groupby('Product')[['Sales']].sum().reset_index()
sales_by_region = df.groupby('Region')[['Sales']].sum().reset_index()
profit_by_product = df.groupby('Product')[['Profit']].sum().reset_index()

# Load workbook and create sheet
wb = load_workbook('Sales_Dashboard.xlsx')
if 'Pivot_Charts' in wb.sheetnames:
    ws = wb['Pivot_Charts']
else:
    ws = wb.create_sheet('Pivot_Charts')

# Write sales_by_product starting at A1
ws.append(['Product', 'Total Sales'])
for r in dataframe_to_rows(sales_by_product, index=False, header=False):
    ws.append(r)

# Determine row where region summary will start
row_start = len(sales_by_product) + 3
ws.cell(row=row_start, column=1).value = 'Region'
ws.cell(row=row_start, column=2).value = 'Total Sales'
for i, r in enumerate(dataframe_to_rows(sales_by_region, index=False, header=False), start=row_start+1):
    ws.append(r)

# Create a bar chart for sales by product
bar = BarChart()
bar.title = 'Sales by Product'
bar.y_axis.title = 'Sales'
bar.x_axis.title = 'Product'
data = Reference(ws, min_col=2, min_row=1, max_row=len(sales_by_product)+1)
cats = Reference(ws, min_col=1, min_row=2, max_row=len(sales_by_product)+1)
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.dataLabels = DataLabelList()
bar.dataLabels.showVal = True
ws.add_chart(bar, 'E2')

# Create a pie chart for sales by region
pie = PieChart()
pie.title = 'Sales by Region'
labels = Reference(ws, min_col=1, min_row=row_start+1, max_row=row_start+len(sales_by_region))
data = Reference(ws, min_col=2, min_row=row_start, max_row=row_start+len(sales_by_region))
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
ws.add_chart(pie, 'E20')

# Save workbook
wb.save('Sales_Dashboard_with_Charts.xlsx')
print('Saved Sales_Dashboard_with_Charts.xlsx')
